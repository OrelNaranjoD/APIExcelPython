from flask import Flask, jsonify, request, redirect
from flask_cors import CORS
from flasgger import Swagger
from flask_jwt_extended import JWTManager, jwt_required, create_access_token
import openpyxl
from dotenv import load_dotenv
import os

# Cargar variables de entorno desde el archivo .env
load_dotenv()

app = Flask(__name__)
CORS(app)

# Configuración de Swagger
swagger_config = {
    "headers": [],
    "specs": [
        {
            "endpoint": "apispec_1",
            "route": "/support/apispec_1.json",
            "rule_filter": lambda rule: True,
            "model_filter": lambda tag: True,
        }
    ],
    "static_url_path": "/support/flasgger_static",
    "swagger_ui": True,
    "specs_route": "/support",
}

swagger_template = {
    "swagger": "2.0",
    "info": {
        "title": "API de Ejemplo con Excel",
        "description": "Esta es una API de ejemplo para manejar usuarios usando un archivo Excel.",
        "version": "1.0.1",
        "license": {"name": "MIT", "url": "https://opensource.org/licenses/MIT"},
    },
    "host": "localhost:5000",
    "basePath": "/",
    "schemes": ["http", "https"],
    "tags": [
        {
            "name": "Usuarios",
            "description": "Operaciones relacionadas con los usuarios",
        },
        {
            "name": "Autenticación",
            "description": "Operaciones relacionadas con la autenticación",
        },
    ],
    "securityDefinitions": {
        "Bearer": {
            "type": "apiKey",
            "name": "Authorization",
            "in": "header",
            "description": 'JWT Authorization header using the Bearer scheme. Example: "Authorization: Bearer {token}"',
        }
    },
    "security": [{"Bearer": []}],
}
swagger = Swagger(app, config=swagger_config, template=swagger_template)

# Configuración de JWT token
app.config["JWT_SECRET_KEY"] = os.getenv("JWT_SECRET_KEY")
jwt = JWTManager(app)


# Función para cargar el archivo Excel
def cargar_excel():
    """Carga el archivo Excel y retorna la hoja activa"""
    try:
        libro = openpyxl.load_workbook("datos.xlsx")
        return libro, libro.active
    except FileNotFoundError:
        return None, None


# Función para buscar un usuario por ID
def buscar_usuario_por_id(hoja, id):
    """Busca un usuario por ID en la hoja de Excel"""
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        id_usuario, nombre, email = fila
        if int(id_usuario) == int(id):
            return {"id": int(id_usuario), "nombre": nombre, "email": email}
    return None


# Función para buscar un usuario por email
def buscar_usuario_por_email(hoja, email):
    """Busca un usuario por email en la hoja de Excel"""
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        id_usuario, nombre, email_usuario = fila
        if email_usuario == email:
            return {"id": int(id_usuario), "nombre": nombre, "email": email_usuario}
    return None


@app.route("/login", methods=["POST"])
def login():
    """
    Solicitar un token JWT
    ---
    tags:
    - Autenticación
    parameters:
      - in: body
        name: usuario
        description: Correo del usuario
        schema:
          type: object
          required:
            - email
          properties:
            email:
              type: string
              description: Email del usuario
              example: "orel.naranjo@prueba.com"
    responses:
      200:
        description: Token JWT generado
        content:
          application/json:
            schema:
              type: object
              properties:
                token:
                  type: string
                  description: El token JWT en formato Bearer
                  example: "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTcyOTExNTE3MywianRpIjoiZDhlYzNkYjQtMTBlYS00YjMxLTg2NmYtMTYzM2RhMDIxYTRiIiwidHlwZSI6ImFjY2VzcyIsInN1YiI6Im9yZWwubmFyYW5qb0BwcnVlYmEuY29tIiwibmJmIjoxNzI5MTE1MTczLCJjc3JmIjoiZDkwNTcwZDktMzc1ZC00NGYxLTk4M2EtOWQ1NTkzMzhmNmZlIiwiZXhwIjoxNzI5MTE2MDczfQ.0_V1qfqAEHCQleSORldg65_oISjwT5idRg0guV04O_U"
      400:
        description: Email es requerido
        content:
          application/json:
            schema:
              type: object
              properties:
                mensaje:
                  type: string
                  example: "Email es requerido"
      401:
        description: Email no encontrado
        content:
          application/json:
            schema:
              type: object
              properties:
                mensaje:
                  type: string
                  example: "Email no encontrado"
      500:
        description: Error al cargar el archivo Excel
        content:
          application/json:
            schema:
              type: object
              properties:
                mensaje:
                  type: string
                  example: "Error: No se pudo cargar el archivo Excel"
    """
    email = request.json.get("email", None)
    if not email:
        return jsonify({"mensaje": "Email es requerido"}), 400

    libro, hoja = cargar_excel()
    if not hoja:
        return jsonify({"mensaje": "Error: No se pudo cargar el archivo Excel"}), 500

    usuario = buscar_usuario_por_email(hoja, email)
    if not usuario:
        return jsonify({"mensaje": "Email no encontrado"}), 401

    access_token = create_access_token(identity=email)
    bearer_token = f"Bearer {access_token}"
    return jsonify(token=bearer_token), 200


# Rutas de Usuarios
@app.route("/usuarios", methods=["GET"])
@jwt_required()
def obtener_usuarios():
    """
    Obtener todos los usuarios
    ---
    tags:
    - Usuarios
    responses:
      200:
        description: Lista de usuarios
        schema:
          type: array
          items:
            type: object
            properties:
              id:
                type: integer
                description: ID del usuario
              nombre:
                type: string
                description: Nombre del usuario
              email:
                type: string
                description: Email del usuario
      401:
        description: "Error: No Autorizado"
        schema:
          type: object
          properties:
            mensaje:
              type: string
              description: Mensaje de error
    """
    usuarios = []
    try:
        libro = openpyxl.load_workbook("datos.xlsx")
        hoja = libro.active
        # Verificar los datos leídos
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            id, nombre, email = fila
            usuarios.append({"id": id, "nombre": nombre, "email": email})

        if not usuarios:
            print("No se encontraron usuarios.")
    except Exception as e:
        print(f"Error al cargar usuarios: {e}")

    return jsonify(usuarios)


@app.route("/usuarios", methods=["POST"])
@jwt_required()
def agregar_usuario():
    """
    Agregar un nuevo usuario
    ---
    tags:
    - Usuarios
    parameters:
      - in: body
        name: usuario
        description: El usuario a agregar
        schema:
          type: object
          required:
            - nombre
            - email
          properties:
            nombre:
              type: string
              description: Nombre del usuario
            email:
              type: string
              description: Email del usuario
    responses:
      201:
        description: Usuario agregado
        schema:
          type: object
          properties:
            mensaje:
              type: string
              description: Mensaje de éxito
            id:
              type: integer
              description: ID del usuario agregado
      409:
        description: "Error: El nombre o el email ya existen"
        schema:
          type: object
          properties:
            mensaje:
              type: string
              description: Mensaje de error
    """
    nuevo_usuario = request.json
    libro, hoja = cargar_excel()
    if not hoja:
        return jsonify({"mensaje": "Error: No se pudo cargar el archivo Excel"}), 500

    # Verificar si el nombre o el email ya existen
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[1] == nuevo_usuario["nombre"] or fila[2] == nuevo_usuario["email"]:
            return jsonify({"mensaje": "Error: El nombre o el email ya existen"}), 409

    # Encontrar el ID más alto y generar uno nuevo
    nuevo_id = (
        max(
            (fila[0] for fila in hoja.iter_rows(min_row=2, values_only=True)), default=0
        )
        + 1
    )

    # Agregar el nuevo usuario
    hoja.append([nuevo_id, nuevo_usuario["nombre"], nuevo_usuario["email"]])
    libro.save("datos.xlsx")
    return jsonify({"mensaje": "Usuario agregado", "id": nuevo_id}), 201


@app.route("/usuarios/<int:id>", methods=["GET"])
@jwt_required()
def obtener_usuario(id):
    """
    Obtener un usuario por ID
    ---
    tags:
    - Usuarios
    parameters:
      - in: path
        name: id
        type: integer
        required: true
        description: ID del usuario
    responses:
      200:
        description: Información del usuario
        schema:
          type: object
          properties:
            id:
              type: integer
              description: ID del usuario
            nombre:
              type: string
              description: Nombre del usuario
            email:
              type: string
              description: Email del usuario
      404:
        description: Usuario no encontrado
    """
    libro, hoja = cargar_excel()
    if not hoja:
        return jsonify({"mensaje": "Error: No se pudo cargar el archivo Excel"}), 500

    usuario = buscar_usuario_por_id(hoja, id)
    if usuario:
        return jsonify(usuario)
    return jsonify({"mensaje": "Usuario no encontrado"}), 404


@app.route("/usuarios/<int:id>", methods=["PUT", "PATCH"])
@jwt_required()
def actualizar_usuario(id):
    """
    Actualizar un usuario existente
    ---
    tags:
    - Usuarios
    parameters:
      - in: path
        name: id
        type: integer
        required: true
        description: ID del usuario
      - in: body
        name: usuario
        description: Datos actualizados del usuario
        schema:
          type: object
          properties:
            nombre:
              type: string
              description: Nombre del usuario
            email:
              type: string
              description: Email del usuario
    responses:
      200:
        description: Usuario actualizado
      404:
        description: Usuario no encontrado
    """
    datos_actualizados = request.json
    libro, hoja = cargar_excel()
    if not hoja:
        return jsonify({"mensaje": "Error: No se pudo cargar el archivo Excel"}), 500

    for fila in hoja.iter_rows(min_row=2):
        if int(fila[0].value) == int(id):
            fila[1].value = datos_actualizados.get("nombre", fila[1].value)
            fila[2].value = datos_actualizados.get("email", fila[2].value)
            libro.save("datos.xlsx")
            return jsonify({"mensaje": "Usuario actualizado"}), 200

    return jsonify({"mensaje": "Usuario no encontrado"}), 404


@app.route("/usuarios/<int:id>", methods=["DELETE"])
@jwt_required()
def eliminar_usuario(id):
    """
    Eliminar un usuario por ID
    ---
    tags:
    - Usuarios
    parameters:
      - in: path
        name: id
        type: integer
        required: true
        description: ID del usuario
    responses:
      200:
        description: Usuario eliminado
      404:
        description: Usuario no encontrado
    """
    libro, hoja = cargar_excel()
    if not hoja:
        return jsonify({"mensaje": "Error: No se pudo cargar el archivo Excel"}), 500

    for fila in hoja.iter_rows(min_row=2):
        if int(fila[0].value) == int(id):
            hoja.delete_rows(fila[0].row)
            libro.save("datos.xlsx")
            return jsonify({"mensaje": "Usuario eliminado"}), 200

    return jsonify({"mensaje": "Usuario no encontrado"}), 404

@app.route('/')
def support_swagger():
    return redirect('/support')

if __name__ == "__main__":
    app.run(debug=True)
